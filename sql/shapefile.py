import geopandas as gpd
import os
import pandas as pd 
from openpyxl import load_workbook
from shapely.geometry import Polygon, MultiPolygon


def converter_shp_para_geojson(pasta_entrada, pasta_saida):
    # Verifica se a pasta de saída existe; se não, cria a pasta
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)
    
    # Itera sobre todos os arquivos na pasta de entrada
    for arquivo in os.listdir(pasta_entrada):
        # Verifica se o arquivo tem a extensão .shp
        if arquivo.endswith('.shp'):
            # Define o caminho completo do arquivo .shp
            caminho_shp = os.path.join(pasta_entrada, arquivo)
            
            # Carrega o shapefile com o geopandas
            gdf = gpd.read_file(caminho_shp)
            
            # Define o nome do arquivo .geojson
            nome_arquivo_geojson = os.path.splitext(arquivo)[0] + '.geojson'
            caminho_geojson = os.path.join(pasta_saida, nome_arquivo_geojson)
            
            # Salva o arquivo como GeoJSON
            gdf.to_file(caminho_geojson, driver="GeoJSON")
            
            print(f"Arquivo convertido: {arquivo} -> {nome_arquivo_geojson}")


def cruza_shp_planilha_perimetro(pasta_saida):
    # Localiza a planilha e o GeoJSON na pasta de saída
    caminho_planilha = None
    caminho_geojson = None
    
    
    for arquivo in os.listdir(pasta_saida):
        if arquivo.endswith('.xlsx') and arquivo.startswith('FAZENDA'):
            caminho_planilha = os.path.join(pasta_saida +'/'+ arquivo)
        elif arquivo.endswith('.geojson') and arquivo.startswith('FAZENDA'):
            caminho_geojson = os.path.join(pasta_saida +'/'+ arquivo)

    # Verifica se ambos os arquivos foram encontrados
    if caminho_planilha and caminho_geojson:
        try:
            # Carrega a planilha e o GeoJSON
            planilha_df = pd.read_excel(caminho_planilha, sheet_name='FAZENDA')
            geo_df = gpd.read_file(caminho_geojson)
            
            geo_df['fazenda']= geo_df['FAZENDA']
            
            # Verifica se as colunas necessárias existem
            if 'cod_fazenda' in planilha_df.columns and 'fazenda' in geo_df.columns:
                # Converte as colunas 'cod_fazenda' e 'fazenda' para o tipo inteiro
                planilha_df['cod_fazenda'] = planilha_df['cod_fazenda'].astype(int)
                
                geo_df['fazenda'] = geo_df['fazenda'].astype(int)
                
                # Faz o merge para obter as colunas 'geometry' e 'area_ha' correspondentes
                merged_df = planilha_df.merge(
                    geo_df[['fazenda', 'geometry', 'area_ha']],
                    left_on='cod_fazenda',
                    right_on='fazenda',
                    how='left'
                )
                
                # Renomeia a coluna 'geometry' para 'geometria'
                merged_df = merged_df.rename(columns={'geometry': 'geometria'})
                
                # Carrega o arquivo Excel existente para edição
                with pd.ExcelWriter(caminho_planilha, mode='a', engine='openpyxl') as writer:
                    # Carrega a planilha existente
                    writer.book = load_workbook(caminho_planilha)
                    
                    # Verifica se a aba 'FAZENDA' já existe antes de escrever nela
                    if 'FAZENDA' in writer.book.sheetnames:
                        # Remove a aba existente antes de adicionar os novos dados
                        std = writer.book['FAZENDA']
                        writer.book.remove(std)
                    
                    # Grava a planilha com as modificações (adicionando as colunas 'geometria' e 'area_ha')
                    merged_df.to_excel(writer, sheet_name='FAZENDA', index=False)
                
                print("Cruzamento e atualização das colunas 'geometria' e 'area_ha' concluídos.")
            else:
                print("A coluna 'cod_fazenda' ou 'fazenda' não foi encontrada em um dos arquivos.")
        except Exception as e:
            print(f"Ocorreu um erro ao processar os arquivos: {e}")
    else:
        print("Arquivo de planilha ou GeoJSON 'FAZENDA' não encontrado em:", pasta_saida)

    
def cruza_shp_planilha_bd(pasta_saida):

    # Localiza a planilha e o GeoJSON na pasta de saída
    caminho_planilha_bd = None
    caminho_geojson_bd = None
    
    
    for arquivo in os.listdir(pasta_saida):
        if arquivo.endswith('.xlsx') and arquivo.startswith('TALHAO'):
            caminho_planilha_bd = os.path.join(pasta_saida +'/'+ arquivo)
        elif arquivo.endswith('.geojson') and arquivo.startswith('TALHAO'):
            caminho_geojson_bd = os.path.join(pasta_saida +'/'+ arquivo)

    # Verifica se ambos os arquivos foram encontrados
    if caminho_planilha_bd and caminho_geojson_bd:
        try:
            # Carrega a planilha e o GeoJSON
            planilha_df = pd.read_excel(caminho_planilha_bd,  sheet_name='TALHAO')
            geo_df = gpd.read_file(caminho_geojson_bd)
            
            # Verifica se as colunas necessárias existem
            if 'CHAVE' in planilha_df.columns and 'CHAVE' in geo_df.columns:
                #enchendo linguisça
                planilha_df['CHAVE'] = planilha_df['CHAVE']
                geo_df['CHAVE'] = geo_df['CHAVE']
                geo_df['area_ha'] = geo_df['AREA_GIS'] #hoje

                # Faz o merge para obter as colunas 'geometry' e 'area_ha' correspondentes
                merged_df = planilha_df.merge(
                    geo_df[['CHAVE', 'geometry', 'area_ha']],
                    left_on='CHAVE',
                    right_on='CHAVE',
                    how='left'
                )
                
                # Renomeia a coluna 'geometry' para 'geometria'
                merged_df = merged_df.rename(columns={'geometry': 'geometria'})
                
                # Carrega o arquivo Excel existente para edição
                with pd.ExcelWriter(caminho_planilha_bd, mode='a', engine='openpyxl') as writer:
                    # Carrega a planilha existente
                    writer.book = load_workbook(caminho_planilha_bd)
                    
                    # Verifica se a aba 'TALHAO' já existe antes de escrever nela
                    if 'TALHAO' in writer.book.sheetnames:
                        # Remove a aba existente antes de adicionar os novos dados
                        std = writer.book['TALHAO']
                        writer.book.remove(std)
                    
                    # Grava a planilha com as modificações (adicionando as colunas 'geometria' e 'area_ha')
                    merged_df.to_excel(writer,  sheet_name='TALHAO', index=False)
                
                print("Cruzamento e atualização das colunas 'geometria' e 'area_ha' concluídos.")
            else:
                print("A coluna 'CHAVE' não foi encontrada em um dos arquivos.")
        except Exception as e:
            print(f"Ocorreu um erro ao processar os arquivos: {e}")
    else:
        print("Arquivo de planilha ou GeoJSON 'TALHAO' não encontrado em:", pasta_saida)


def talhao_fazenda(caminho):
    
    """
    Processa shapefiles no diretório especificado: dissolve pela coluna 'FAZENDA',
    aplica buffer de 10m, simplifica as geometrias (tolerância de 5m) e remove buracos.
    Salva os arquivos resultantes no mesmo diretório com o prefixo 'FAZENDA_'.

    Args:
        caminho (str): Diretório contendo os arquivos shapefile.
    """
    for arquivo in os.listdir(caminho):
        if arquivo.startswith('TALHAO') and arquivo.endswith('.shp'):
            try:
                caminho_arquivo = os.path.join(caminho, arquivo)
                # Lendo o shapefile
                gdf = gpd.read_file(caminho_arquivo)
                
                if 'FAZENDA' not in gdf.columns:
                    print(f"A coluna 'FAZENDA' não foi encontrada em {arquivo}.")
                    continue
                
                # Dissolver pela coluna 'FAZENDA'
                gdf_dissolve = gdf.dissolve(by='FAZENDA')
                
                # Projeção temporária para métrico
                utm_crs = gdf_dissolve.estimate_utm_crs()
                gdf_dissolve = gdf_dissolve.to_crs(utm_crs)
                
                # Aplicar buffer de 10 metros
                gdf_dissolve['geometry'] = gdf_dissolve.buffer(10)
                
                # Simplificar geometrias (reduzir vértices) em 5 metros
                gdf_dissolve['geometry'] = gdf_dissolve.simplify(tolerance=5)
                
                # Remover buracos das geometrias
                def processar_geometria(geom):
                    if geom.is_empty or geom is None:
                        return geom
                    elif isinstance(geom, Polygon):
                        return Polygon(geom.exterior)  # Remove buracos de Polygon
                    elif isinstance(geom, MultiPolygon):
                        # Remove buracos de cada Polygon dentro do MultiPolygon
                        return MultiPolygon([Polygon(p.exterior) for p in geom.geoms])
                    else:
                        return geom

                gdf_dissolve['geometry'] = gdf_dissolve['geometry'].apply(processar_geometria)

                gdf_dissolve['area_ha'] = gdf_dissolve['geometry'].area / 10000
                
                # Reprojetar para o CRS original (graus)
                gdf_dissolve = gdf_dissolve.to_crs(gdf.crs)
                
                # Nome do arquivo de saída
                nome_saida = f"FAZENDA_{arquivo}"
                caminho_saida = os.path.join(caminho, nome_saida)
                
                # Salvando o shapefile processado
                gdf_dissolve.to_file(caminho_saida)
                print(f"Arquivo salvo com sucesso: {caminho_saida}")
            
            except Exception as e:
                print(f"Erro ao processar {arquivo}: {e}")

